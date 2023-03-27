import sys
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QMenu, QComboBox, QStyledItemDelegate
import openpyxl
from PyQt5.QtGui import QColor


class dataTypeDelegate(QStyledItemDelegate):
    ''' Delegate for the data type columns '''
    ''' This is used to create a combobox in the table '''

    def createEditor(self, parent, option, index):
        editor = QComboBox(parent)
        # add datatypes available for siemens tia portal v16 plc tags
        siemens_data_types = [
            "Bool",    # Boolean (1 bit)
            "Byte",    # Byte (8 bits)
            "Word",    # Word (16 bits)
            "DWord",   # Double Word (32 bits)
            "LWord",   # Long Word (64 bits)
            "SInt",    # Signed 8-bit integer
            "Int",     # Signed 16-bit integer
            "DInt",    # Signed 32-bit integer
            "LInt",    # Signed 64-bit integer
            "USInt",   # Unsigned 8-bit integer
            "UInt",    # Unsigned 16-bit integer
            "UDInt",   # Unsigned 32-bit integer
            "ULInt",   # Unsigned 64-bit integer
            "Real",    # Floating-point number (32 bits)
            "LReal",   # Floating-point number (64 bits)
            "Time",    # Time duration (32 bits)
            "LTime",   # Time duration (64 bits)
            "Date",    # Calendar date (16 bits)
            "TIME_OF_DAY", # Time of day (32 bits)
            "DATE_AND_TIME", # Combined date and time (64 bits)
            "Char",    # Single ASCII character (8 bits)
            "WChar",   # Single Unicode character (16 bits)
            "STring",  # Variable-length ASCII string
            "WSTring", # Variable-length Unicode string
            "Array",   # Array of specified data type
            "Struct",  # Structure containing multiple data types
        ]
        editor.addItems(siemens_data_types)  # Add available data types
        return editor

    def setEditorData(self, editor, index):
        value = index.model().data(index, Qt.EditRole)

        if isinstance(editor, QComboBox):
            editor.setCurrentText(value)
        else:
            editor.setText(value)

    def setModelData(self, editor, model, index):
        if isinstance(editor, QComboBox):
            value = editor.currentText()
        else:
            value = editor.text()

        model.setData(index, value, Qt.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)

class ExcelImporter(QMainWindow):
    def __init__(self, file_path):
        super().__init__()

        self.file_path = file_path
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Excel Importer")

        # Create the central widget and layout
        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        # Load the Excel workbook
        self.workbook = openpyxl.load_workbook(self.file_path)

        # Add a QTableWidget for each sheet in the workbook
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            table_widget = self.create_table_widget(sheet)
            self.layout.addWidget(table_widget)

    def create_table_widget(self, sheet):
        # Create a QTableWidget with the same number of rows and columns as the sheet
        header_labels = [str(cell.value) for cell in sheet[1]]
        table_widget = QTableWidget(sheet.max_row, sheet.max_column)
        table_widget.setHorizontalHeaderLabels(header_labels)
        table_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        table_widget.customContextMenuRequested.connect(self.show_context_menu)
        table_widget.setAlternatingRowColors(True)
        # table_widget.setItemDelegate(CustomDelegate())
        # Find 'Data Type' column index and set the custom delegate for it
        for col_num, header_label in enumerate(header_labels):
            if header_label == "Data Type":
                table_widget.setItemDelegateForColumn(col_num, dataTypeDelegate())


        # Populate the table with data from the sheet
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=1):
            for col_num, cell in enumerate(row, start=0):
                table_widget.setItem(row_num - 1, col_num, QTableWidgetItem(str(cell.value)))

        return table_widget

    def show_context_menu(self, pos):
        menu = QMenu(self)
        menu.addAction("Action 1", self.action1)
        menu.addAction("Action 2", self.action2)
        menu.exec_(self.mapToGlobal(pos))   

    def action1(self):
        print("Action 1 triggered") 

    def action2(self):
        print("Action 2 triggered")



if __name__ == "__main__":
    app = QApplication(sys.argv)
    file_path = "PLCTags.xlsx"  # Replace this with the path to your .xlsx file
    main_window = ExcelImporter(file_path)
    main_window.show()
    sys.exit(app.exec_())
